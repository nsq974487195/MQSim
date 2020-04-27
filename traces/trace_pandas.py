#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import pandas as pd

import matplotlib.pyplot as plt


def writeexcel(maptable, tracename, i):

    wb = Workbook()

    dest_filename = 'trace_analysis.xlsx'

    if(i == 0):
        ws = wb.active

        ws.title = tracename

    else:

        ws = wb.create_sheet(title=tracename)

    for col in range(1, len(maptable) + 1):

        ws.cell(column=col, row=1, value=col - 1)

        ws.cell(column=col, row=2, value=maptable[col - 1]['update_count'])

        ws.cell(column=col, row=3, value=maptable[col - 1]['update_avg'])

        ws.cell(column=col, row=4, value=maptable[col - 1]['update_min'])

        ws.cell(column=col, row=5, value=maptable[col - 1]['update_max'])

    wb.save(filename=dest_filename)


# 统计写更新的操作次数和写更新的时间间隔

def updateinfo(tracename):

    file = open(tracename)

    page_size = 4

    update_count = 0

    entry = {"lpn": 0, "access_time": 0.0, "update_time": 0.0, "update_count": 0, "update_min": 0, "update_max": 0}
    

    large_lpn = 0

    capacity = 32 * 1024 * 1024  # 128G SSD = 32 * 1024* 1024*4KB
    
    #maptable = pd.DataFrame([ [0.0,0.0,0,0.0,0.0,0.0,0] for i in range(0, capacity)],columns=['access_time', 'update_interval', 'update_count', 'update_avg', "update_min", "update_max", "write_flag"])
    maptable = [{"access_time": 0.0, "update_interval": 0.0, "update_count": 0, "update_avg": 0.0, "update_min": 1.0 * 10**12, "update_max": 0.0, "write_flag": 0} for i in range(0, capacity)]  # write_flag =1 表示已经被写

    print("初始化完成 maptable:", len(maptable), sys.getsizeof(maptable) / (1024 * 1024))


    #print("初始化完成 maptable:", maptable.index,maptable.columns, sys.getsizeof(maptable) / (1024 * 1024))
    
    
    #trace df
    #tracedf=pd.read_table(tracename,sep=" ",header=None)

    #file = open(tracename)

    while 1:
        lines = file.readlines(100000)
        if not lines:
            break
        for line in lines:

            tmp = line.split()

            #print(tmp)
  

            if(int(tmp[4]) == 1):  # read

                first_lpn = int(int(tmp[2]) / page_size) % capacity

                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):

                    #print("first_lpn", first_lpn)

                    if(maptable[first_lpn]['write_flag'] == 1):  # 被写过

                        # maptable[int(int(tmp[2]) / page_size)]['access_time']=float(tmp[0]) #本次访问的时间

                        maptable[first_lpn]['update_interval'] = float(tmp[0]) - maptable[first_lpn]['access_time']  # 本次访问的时间

                        maptable[first_lpn]['update_count'] = maptable[first_lpn]['update_count'] + 1

                        # if(maptable[first_lpn]['update_interval'] <= 10):

                        #     print(tmp)

                        if(maptable[first_lpn]['update_min'] > maptable[first_lpn]['update_interval']):

                            maptable[first_lpn]['update_min'] = maptable[first_lpn]['update_interval']

                        if(maptable[first_lpn]['update_max'] < maptable[first_lpn]['update_interval']):

                            maptable[first_lpn]['update_max'] = maptable[first_lpn]['update_interval']

                        maptable[first_lpn]['update_avg'] = maptable[first_lpn]['update_avg'] + (maptable[first_lpn]['update_interval'] - maptable[first_lpn]['update_avg']) / maptable[first_lpn]['update_count']

                    first_lpn = first_lpn + 1

            elif(int(tmp[4]) == 0):

                first_lpn = int(int(tmp[2]) / page_size) % capacity

                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):

                    maptable[first_lpn]['write_flag'] = 1

                    maptable[first_lpn]['access_time'] = float(tmp[0])

                    first_lpn = first_lpn + 1

                
    
    #for i in range(0, len(tracedf)):
        
        #tmp=tracedf.loc[i].tolist()#list(tracedf.loc[i])
        
        #print(tmp)
        
#            if(int(tmp[4]) == 1):  # read
#    
#                first_lpn = int(int(tmp[2]) / page_size)% capacity
#    
#                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):
#    
#                    #print("first_lpn", first_lpn)
#    
#                    if(maptable.loc[first_lpn,'write_flag'] == 1):  # 被写过
#    
#                        # maptable[int(int(tmp[2]) / page_size)]['access_time']=float(tmp[0]) #本次访问的时间
#    
#                        maptable.loc[first_lpn,'update_interval'] = float(tmp[0]) - maptable.loc[first_lpn,'access_time']  # 本次访问的时间
#    
#                        maptable.loc[first_lpn,'update_count'] = maptable.loc[first_lpn,'update_count'] + 1
#    
#                        # if(maptable[first_lpn]['update_interval'] <= 10):
#    
#                        #     print(tmp)
#    
#                        if(maptable.loc[first_lpn,'update_min'] > maptable.loc[first_lpn,'update_interval']):
#    
#                            maptable.loc[first_lpn,'update_min'] = maptable.loc[first_lpn,'update_interval']
#    
#                        if(maptable.loc[first_lpn,'update_max'] < maptable.loc[first_lpn,'update_interval']):
#    
#                            maptable.loc[first_lpn,'update_max'] = maptable.loc[first_lpn,'update_interval']
#    
#                        maptable.loc[first_lpn,'update_avg'] = maptable.loc[first_lpn,'update_avg'] + (maptable.loc[first_lpn,'update_interval'] - maptable.loc[first_lpn,'update_avg']) / maptable.loc[first_lpn,'update_count']
#    
#                    first_lpn = first_lpn + 1
#    
#            elif(int(tmp[4]) == 0):
#    
#                first_lpn = int(int(tmp[2]) / page_size)% capacity
#    
#                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):
#    
#                    maptable.loc[first_lpn,'write_flag'] = 1
#    
#                    maptable.loc[first_lpn,'access_time'] = float(tmp[0])
#    
#                    first_lpn = first_lpn + 1
#
#    update_min = maptable.loc[0,"update_min"]

    update_max = 0

    update_interval = 0

    update_count = 0

    update_avg = 0

#    for i in range(0, len(maptable)):
#
#        if(maptable.loc[i,"update_avg"] > update_avg):
#
#            update_avg = maptable.loc[i,"update_avg"]
#
#        if(maptable.loc[i,"update_min"] < update_min):
#
#            update_min = maptable.loc[i,"update_min"]
#
#        if(maptable.loc[i,"update_max"] > update_max):
#
#            update_max = maptable.loc[i,"update_max"]
    
    maptable=pd.DataFrame(maptable,columns=['access_time', 'update_interval', 'update_count', 'update_avg', "update_min", "update_max", "write_flag"])

    print("trace:%s update_avg:%fns, update_avg / (10 ** 6):%fs,update_min:%fns,update_max:%fns" % (tracename, maptable['update_avg'].mean(), maptable['update_avg'].mean() / (10 ** 6), maptable['update_min'].mean(), maptable['update_max'].mean()))

#    writeexcel(maptable, tracename, 0)
    
    maptable.sort_values("update_avg",inplace=True)
    
    print("<0的百分比:%f >0和<0的百分比:%f "%(len(maptable[(maptable['update_avg']==0.0)].index)/len(maptable.index),len(maptable[(maptable['update_avg']>0.0) & (maptable['update_avg']<100.0*10**6)].index)/len(maptable.index)))
    
    updatetable=maptable[(maptable['update_avg']>0.0)]
    
    plt.figure()
    
    updatetable['update_avg'].plot(kind='kde',style='k--')
    #updatetable['update_avg'].hist( cumulative = True )

    maptable = []






if __name__ == '__main__':
    # max()
    # export_result()

    tracename = ["HM_0", "HM_1", "PRN_0", "PRN_1", "PROJ_1", "PROJ_3", "PROJ_4", "PRXY_0", "RSRCH", "SRC1_2", "SRC2_0", "SRC2_2", "STG_0", "USR_0", "WDEV_0"]

    # for item in tracename:

    # average(item)

    updateinfo("HM_0")


