#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def writeexcel(maptable, tracename, i):

    wb = Workbook()

    dest_filename = 'trace_analysis.xlsx'

    if(i == 0):
        ws = wb.active

        ws.title = tracename

    else:

        ws = wb.create_sheet(title=tracename)

    for col in range(1, len(maptable) + 1):

        ws.cell(column=col, row=1, value=col - 1)

        ws.cell(column=col, row=2, value=maptable[col - 1]['update_count'])

        ws.cell(column=col, row=3, value=maptable[col - 1]['update_avg'])

        ws.cell(column=col, row=4, value=maptable[col - 1]['update_min'])

        ws.cell(column=col, row=5, value=maptable[col - 1]['update_max'])

    wb.save(filename=dest_filename)


# 统计写更新的操作次数和写更新的时间间隔


def updateinfo(tracename):

    file = open(tracename)

    page_size = 4

    update_count = 0

    entry = {"lpn": 0, "access_time": 0.0, "update_time": 0.0, "update_count": 0, "update_min": 0, "update_max": 0}

    large_lpn = 0

    capacity = 32 * 1024 * 1024  # 128G SSD = 32 * 1024* 1024*4KB

    # while 1:
    #     lines = file.readlines(100000)
    #     if not lines:
    #         break
    #     for line in lines:

    #         tmp = line.split()

    #         if(int(int(tmp[2]) / page_size) + int(tmp[3]) / page_size > large_lpn):

    #             large_lpn = int(int(tmp[2]) / page_size + int(tmp[3]) / page_size)

    # print("large_lpn:%d" % large_lpn)

    maptable = [{"access_time": 0.0, "update_interval": 0.0, "update_count": 0, "update_avg": 0.0, "update_min": 1.0 * 10**12, "update_max": 0.0, "write_flag": 0} for i in range(0, capacity)]  # write_flag =1 表示已经被写

    print("初始化完成 maptable:", len(maptable), sys.getsizeof(maptable) / (1024 * 1024))

    file = open(tracename)

    while 1:
        lines = file.readlines(10000)
        if not lines:
            break
        for line in lines:

            tmp = line.split()

            # print(tmp)

            if(int(tmp[4]) == 1):  # read

                first_lpn = int(int(tmp[2]) / page_size)

                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):

                    #print("first_lpn", first_lpn)

                    if(maptable[first_lpn]['write_flag'] == 1):  # 被写过

                        # maptable[int(int(tmp[2]) / page_size)]['access_time']=float(tmp[0]) #本次访问的时间

                        maptable[first_lpn]['update_interval'] = float(tmp[0]) - maptable[first_lpn]['access_time']  # 本次访问的时间

                        maptable[first_lpn]['update_count'] = maptable[first_lpn]['update_count'] + 1

                        # if(maptable[first_lpn]['update_interval'] <= 10):

                        #     print(tmp)

                        if(maptable[first_lpn]['update_min'] > maptable[first_lpn]['update_interval']):

                            maptable[first_lpn]['update_min'] = maptable[first_lpn]['update_interval']

                        if(maptable[first_lpn]['update_max'] < maptable[first_lpn]['update_interval']):

                            maptable[first_lpn]['update_max'] = maptable[first_lpn]['update_interval']

                        maptable[first_lpn]['update_avg'] = maptable[first_lpn]['update_avg'] + (maptable[first_lpn]['update_interval'] - maptable[first_lpn]['update_avg']) / maptable[first_lpn]['update_count']

                    first_lpn = first_lpn + 1

            elif(int(tmp[4]) == 0):

                first_lpn = int(int(tmp[2]) / page_size)

                while(first_lpn <= int(int(tmp[2]) / page_size + int(tmp[3]) / page_size) % capacity):

                    maptable[first_lpn]['write_flag'] = 1

                    maptable[first_lpn]['access_time'] = float(tmp[0])

                    first_lpn = first_lpn + 1

    update_min = maptable[0]["update_min"]

    update_max = 0

    update_interval = 0

    update_count = 0

    update_avg = 0

    # for item in maptable:

    #     if(item["update_avg"] > update_avg):

    #         update_avg = item["update_avg"]

    #     if(item["update_min"] < update_min):

    #         update_min = item["update_min"]

    #     if(item["update_max"] > update_max):

    #         update_max = item["update_max"]

    # print("trace:%s update_avg:%fns, update_avg / (10 ** 6):%fs,update_min:%fns,update_max:%fns" % (tracename, update_avg, update_avg / (10 ** 6), update_min, update_max))

#    writeexcel(maptable, tracename, 0)

    maptable = []


def average(tracename):

    file = open(tracename)

    count = 0

    read_count = 0

    read_avg = 0

    read_small_4k = 0

    read_small_8k = 0

    read_small_16k = 0

    read_small_32k = 0

    write_count = 0

    write_avg = 0

    write_small_4k = 0

    write_small_8k = 0

    write_small_16k = 0

    write_small_32k = 0

    while 1:
        lines = file.readlines(100000)
        if not lines:
            break
        for line in lines:

            tmp = line.split()

            count = count + 1

            if(int(tmp[4]) == 1):  # read

                read_count = read_count + 1

                read_avg = read_avg + float((int(tmp[3]) / 2 - read_avg) / read_count)

                if((int(tmp[3]) / 2) <= 4):

                    read_small_4k = read_small_4k + 1  # + float((int(tmp[3]) / 2 - read_small) / read_count)

                elif((int(tmp[3]) / 2) <= 8):

                    read_small_8k = read_small_8k + 1

                elif((int(tmp[3]) / 2) <= 16):

                    read_small_16k = read_small_16k + 1
                else:

                    read_small_32k = read_small_32k + 1

            elif(int(tmp[4]) == 0):

                write_count = write_count + 1

                write_avg = write_avg + float((int(tmp[3]) / 2 - write_avg) / write_count)

                if((int(tmp[3]) / 2) <= 4):

                    write_small_4k = write_small_4k + 1  # + float((int(tmp[3]) / 2 - write_small) / write_count)
                elif((int(tmp[3]) / 2) <= 8):
                    write_small_8k = write_small_8k + 1

                elif((int(tmp[3]) / 2) <= 16):
                    write_small_16k = write_small_16k + 1

                else:
                    write_small_32k = write_small_32k + 1

    # print("trace:%s read:%f write:%f  read_avg:%f read_small:%f  write_avg:%f write_small:%f " % (tracename, read_count / count, write_count / count, read_avg, read_small / count, write_avg, write_small / count))
    #print("trace   read   avg   4k  8k  16k  32k  avg  4k  8k  16k  32k")
    #print("%s &%.2f &%.2f &%.2f &%.2f &%.2f &%.2f &%.2f & %.2f &%.2f &%.2f &%.2f &%.2f\\\ \\hline" % (tracename, read_count / count, write_count / count, read_avg, read_small_4k / count, read_small_8k / read_count, read_small_16k / read_count, read_small_32k / read_count, write_avg, write_small_4k / write_count, write_small_8k / write_count, write_small_16k / write_count, write_small_32k / write_count))
    print("%s  &%.2f &%.2f &%.2f &%.2f &%.2f &%.2f & %.2f\\\ \\hline" % (tracename, read_count / count, read_avg, write_avg, (read_small_4k + write_small_4k) / count, (read_small_8k + write_small_8k) / count, (read_small_16k + write_small_16k) / count, (read_small_32k + write_small_32k) / count))


def max():

    file = open("HM_0_out")
    lineno = 0
    latency = 0
    count = 0
    while 1:
        lines = file.readlines(100000)
        if not lines:
            break
        for line in lines:
            lineno = lineno + 1
            if (lineno >= 168) and (lineno <= 1009500):
                result = line.split()
               # print(result)
                if(latency < int(result[6])) and (int(result[3]) == 2045):
                    latency = int(result[6])
                    no = lineno
                if(int(result[6]) >= 999999) and(int(result[3]) == 2045):
                    count = count + 1

    print("max:%d no.:%d count:%d" % (latency, no, count))


def export_result():

    tracename = ["HM_0", "HM_1", "PRN_0", "PRN_1", "PROJ_1", "PROJ_3", "PROJ_4", "PRXY_0", "RSRCH", "SRC1_2", "SRC2_0", "SRC2_2", "STG_0", "USR_0", "WDEV_0"]

    perform_file = open("perf.txt", 'w')

    pageno_file = open("pageno.txt", 'w')

    delete_file = open("del.txt", 'w')

    for item in tracename:

        filename = item + "_st"

        file = open(filename)

        while 1:

            lines = file.readlines(10000)

            if not lines:

                break
            for line in lines:

                if (line.startswith("delete request average response time:")):

                    tmp = line.split()

                    print(tmp[-1])

                    delete_file.write(tmp[-1] + '\n')

                if (line.startswith("write request average response time:")):

                    tmp = line.split()

                    print(tmp[-1])

                    perform_file.write(tmp[-1] + '\n')

                if (line.startswith("the NO. of page readed by secure deletion:")):

                    tmp = line.split()

                    print(tmp[-1])

                    pageno_file.write(tmp[-1] + '\n')


if __name__ == '__main__':
    # max()
    # export_result()

    tracename = ["HM_0", "HM_1", "PRN_0", "PRN_1", "PROJ_1", "PROJ_3", "PROJ_4", "PRXY_0", "RSRCH", "SRC1_2", "SRC2_0", "SRC2_2", "STG_0", "USR_0", "WDEV_0"]

    for item in tracename:

        average(item)

    # updateinfo("HM_0")
